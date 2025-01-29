import logging
import openpyxl
import csv
import os
from typing import Dict, Any, List


logger = logging.getLogger(__name__)


class Database:
    def __init__(self) -> None:
        self.format_types = ["csv", "xlsx"]

    def write(self, file: str, data: List[Dict[str, Any]], mode: str = "a") -> None:
        if not file or not isinstance(file, str):
            raise ValueError("Invalid file name")
        if not data or not isinstance(data, list):
            raise ValueError("Data must be a non-empty list")
        if mode not in {"a", "w"}:
            raise ValueError("Mode must be 'a' (append) or 'w' (write)")

        filename, format_type = file.strip().rsplit(".", 1)
        if format_type not in self.format_types:
            raise ValueError(f"Unsupported file format: {format_type}")
        if not filename:
            raise ValueError("File name cannot be empty")

        try:
            if format_type == "csv":
                self._write_csv(f"{filename}.csv", data, mode)
            elif format_type == "xlsx":
                self._write_excel(f"{filename}.xlsx", data, mode)

        except Exception as e:
            logger.exception(f"Failed to write data to {file}: {e}")
            raise

    def _write_csv(self, file: str, data: List[Dict[str, Any]], mode: str = "a") -> None:
        write_header = mode == "w" or not os.path.exists(file)

        with open(file, mode=mode, newline="") as csvfile:
            writer = csv.writer(csvfile)

            if write_header or not csvfile.tell():
                writer.writerow(data[0].keys())

            for row in data:
                writer.writerow(row.values())

    def _write_excel(self, file: str, data: List[Dict[str, Any]], mode: str = "a") -> None:
        if mode == "a" and os.path.exists(file):
            wb = openpyxl.load_workbook(file)
        else:
            wb = openpyxl.Workbook()

        ws = wb.active
        if ws.max_row == 1 and not ws.cell(row=1, column=1).value:
            for index, header in enumerate(data[0], start=1):
                ws.cell(row=1, column=index, value=header)

        for item in data:
            ws.append(list(item.values()))

        wb.save(file)
